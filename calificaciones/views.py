import io
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie

from .models import (
    AlumnoCalificacion,
    CategoriaEvaluacion,
    NotaEvaluacion,
    BoletinConfig,
)


def menu(request):
    return render(request, "calificaciones/menu.html")


def _materias_existentes():
    return (
        AlumnoCalificacion.objects.values_list("materia", flat=True)
        .distinct()
        .order_by("materia")
    )


def _alumnos_por_materia(materia):
    return list(
        AlumnoCalificacion.objects.filter(materia=materia)
        .order_by("nombre")
        .values("id", "nombre", "curso", "division")
    )


def _categorias_por_materia(materia):
    return list(
        CategoriaEvaluacion.objects.filter(materia_nombre=materia)
        .order_by("orden", "nombre")
        .values("id", "nombre", "orden")
    )


def _notas_para_materia(materia, cuat):
    qs = NotaEvaluacion.objects.filter(
        alumno_calificacion__materia=materia,
        cuatrimestre=cuat,
    ).select_related("alumno_calificacion", "categoria")

    notas = {}
    for n in qs:
        key = f"{n.alumno_calificacion_id}_{n.categoria_id}"
        if key not in notas:
            notas[key] = []
        notas[key].append({
            "id": n.id,
            "valor": str(n.valor),
            "descripcion": n.descripcion or "",
        })
    return notas


def _redondear(valor, decimales=2):
    """Redondea un Decimal a la cantidad de decimales indicada (mitad hacia arriba)."""
    q = Decimal("1").scaleb(-decimales)
    return valor.quantize(q, rounding=ROUND_HALF_UP)


def _promedio_alumno_cuat(alumno_id, cuat):
    notas = NotaEvaluacion.objects.filter(
        alumno_calificacion_id=alumno_id,
        cuatrimestre=cuat,
    )
    if not notas.exists():
        return None
    total = sum((n.valor for n in notas), Decimal("0"))
    promedio = total / Decimal(len(notas))
    return _redondear(promedio)


def _promedio_materia_cuat(materia, cuat):
    alumnos = _alumnos_por_materia(materia)
    promedios = []
    for a in alumnos:
        p = _promedio_alumno_cuat(a["id"], cuat)
        if p is not None:
            promedios.append(p)
    if not promedios:
        return None
    total = sum(promedios, Decimal("0"))
    promedio = total / Decimal(len(promedios))
    return _redondear(promedio)


def _estado(nota):
    if nota is None:
        return "-"
    return "Aprobado" if nota >= 6 else "Desaprobado"


# ─────────────────────────────────────────────
#  ABM DE ALUMNOS (por materia)
# ─────────────────────────────────────────────

def profesor_listar_alumnos(request, materia):
    alumnos = list(
        AlumnoCalificacion.objects.filter(materia=materia)
        .order_by("apellido", "nombre")
        .values("id", "apellido", "nombre", "dni", "curso", "division")
    )
    return render(request, "calificaciones/gestionar_alumnos.html", {
        "materia": materia,
        "alumnos": alumnos,
    })


def profesor_agregar_alumno(request, materia):
    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        dni = request.POST.get("dni", "").strip()
        curso = request.POST.get("curso", "").strip()
        division = request.POST.get("division", "").strip()

        if not nombre:
            messages.error(request, "El nombre es obligatorio.")
        elif not apellido:
            messages.error(request, "El apellido es obligatorio.")
        elif dni:
            existe = AlumnoCalificacion.objects.filter(
                dni__iexact=dni, materia=materia
            ).exclude(dni="").exists()
            if existe:
                messages.error(request, f"Ya existe un alumno con DNI {dni} en {materia}.")
            else:
                AlumnoCalificacion.objects.create(
                    nombre=nombre, apellido=apellido, dni=dni,
                    curso=curso or "-", division=division or "-", materia=materia,
                )
                messages.success(request, f"Alumno {apellido}, {nombre} cargado.")
        else:
            AlumnoCalificacion.objects.create(
                nombre=nombre, apellido=apellido, dni="",
                curso=curso or "-", division=division or "-", materia=materia,
            )
            messages.success(request, f"Alumno {apellido}, {nombre} cargado.")

    return redirect("profesor_listar_alumnos", materia=materia)


def profesor_editar_alumno(request, materia, alumno_id):
    alumno = get_object_or_404(AlumnoCalificacion, id=alumno_id, materia=materia)

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        apellido = request.POST.get("apellido", "").strip()
        dni = request.POST.get("dni", "").strip()
        curso = request.POST.get("curso", "").strip()
        division = request.POST.get("division", "").strip()

        if not nombre or not apellido:
            messages.error(request, "Nombre y apellido son obligatorios.")
        else:
            if dni:
                existe = (
                    AlumnoCalificacion.objects
                    .filter(dni__iexact=dni, materia=materia)
                    .exclude(id=alumno.id)
                    .exclude(dni="")
                    .exists()
                )
                if existe:
                    messages.error(request, f"Ya existe otro alumno con DNI {dni} en {materia}.")
                    return redirect("profesor_editar_alumno", materia=materia, alumno_id=alumno.id)
            alumno.nombre = nombre
            alumno.apellido = apellido
            alumno.dni = dni
            alumno.curso = curso or "-"
            alumno.division = division or "-"
            alumno.save()
            messages.success(request, "Alumno actualizado.")
            return redirect("profesor_listar_alumnos", materia=materia)

    return render(request, "calificaciones/editar_alumno.html", {
        "materia": materia,
        "alumno": alumno,
    })


def profesor_eliminar_alumno(request, materia, alumno_id):
    alumno = get_object_or_404(AlumnoCalificacion, id=alumno_id, materia=materia)
    alumno.delete()
    messages.success(request, "Alumno eliminado.")
    return redirect("profesor_listar_alumnos", materia=materia)


# ─────────────────────────────────────────────
#  PROFESOR
# ─────────────────────────────────────────────

def profesor_seleccionar_materia(request):
    materias = _materias_existentes()
    return render(request, "calificaciones/profesor_materias.html", {
        "materias": materias,
    })


@ensure_csrf_cookie
def profesor_ver_alumnos(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    alumnos = _alumnos_por_materia(materia)
    categorias = _categorias_por_materia(materia)
    notas = _notas_para_materia(materia, cuat)

    for a in alumnos:
        a["celdas"] = []
        total = Decimal("0")
        count = 0
        for cat in categorias:
            key = f"{a['id']}_{cat['id']}"
            cat_notas = notas.get(key, [])
            valor = cat_notas[-1]["valor"] if cat_notas else ""
            a["celdas"].append({
                "cat_id": cat["id"],
                "valor": valor,
            })
            for n in cat_notas:
                try:
                    total += Decimal(n["valor"])
                    count += 1
                except (InvalidOperation, TypeError):
                    pass
        a["promedio"] = str(_redondear(total / Decimal(count))) if count > 0 else "-"
        a["estado"] = _estado(total / Decimal(count) if count > 0 else None)

    boletin = BoletinConfig.objects.filter(cuatrimestre=cuat).first()

    return render(request, "calificaciones/profesor_notas.html", {
        "cuat": cuat,
        "cuat_display": f"{cuat}° Cuatrimestre",
        "materia": materia,
        "alumnos": alumnos,
        "categorias": categorias,
        "boletin": boletin,
    })


def profesor_guardar_nota(request, cuat, materia):
    """Endpoint AJAX: guarda la nota de una celda en particular."""
    if cuat not in (1, 2):
        return JsonResponse({"ok": False, "error": "Cuatrimestre inválido"}, status=400)

    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método no permitido"}, status=405)

    alumno_id = request.POST.get("alumno_id")
    cat_id = request.POST.get("cat_id")
    valor = request.POST.get("valor", "").strip()

    if not alumno_id or not cat_id:
        return JsonResponse({"ok": False, "error": "Faltan datos"}, status=400)

    alumno = AlumnoCalificacion.objects.filter(id=alumno_id, materia=materia).first()
    categoria = CategoriaEvaluacion.objects.filter(id=cat_id, materia_nombre=materia).first()
    if not alumno or not categoria:
        return JsonResponse({"ok": False, "error": "Alumno o categoría no encontrados"}, status=404)

    if valor == "":
        # celda vacía => borrar nota
        NotaEvaluacion.objects.filter(
            alumno_calificacion=alumno,
            categoria=categoria,
            cuatrimestre=cuat,
        ).delete()
        return JsonResponse({"ok": True, "borrado": True})

    try:
        valor_dec = Decimal(valor)
    except InvalidOperation:
        return JsonResponse({"ok": False, "error": f"'{valor}' no es una nota válida"}, status=400)

    if valor_dec < 0 or valor_dec > 10:
        return JsonResponse({"ok": False, "error": "La nota debe estar entre 0 y 10"}, status=400)

    NotaEvaluacion.objects.update_or_create(
        alumno_calificacion=alumno,
        categoria=categoria,
        cuatrimestre=cuat,
        defaults={"valor": valor_dec},
    )

    # Recalcular promedio del alumno
    prom = _promedio_alumno_cuat(alumno_id, cuat)
    return JsonResponse({
        "ok": True,
        "promedio": str(prom) if prom is not None else "-",
        "estado": _estado(prom),
    })


def profesor_categorias(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    cats = _categorias_por_materia(materia)
    return render(request, "calificaciones/gestionar_categorias.html", {
        "cuat": cuat,
        "materia": materia,
        "categorias": cats,
    })


def profesor_agregar_categoria(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        if nombre:
            max_orden = (
                CategoriaEvaluacion.objects
                .filter(materia_nombre=materia)
                .values_list("orden", flat=True)
            )
            orden = (max(max_orden, default=0)) + 1 if max_orden else 1
            CategoriaEvaluacion.objects.create(
                nombre=nombre,
                materia_nombre=materia,
                orden=orden,
            )
            messages.success(request, f"Categoría '{nombre}' creada.")
        else:
            messages.error(request, "El nombre no puede estar vacío.")

    return redirect("profesor_categorias", cuat=cuat, materia=materia)


def profesor_editar_categoria(request, cuat, materia, cat_id):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    cat = get_object_or_404(
        CategoriaEvaluacion, id=cat_id, materia_nombre=materia
    )

    if request.method == "POST":
        nombre = request.POST.get("nombre", "").strip()
        if nombre:
            cat.nombre = nombre
            cat.save()
            messages.success(request, "Categoría actualizada.")
        else:
            messages.error(request, "El nombre no puede estar vacío.")

    return redirect("profesor_categorias", cuat=cuat, materia=materia)


def profesor_eliminar_categoria(request, cuat, materia, cat_id):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    cat = get_object_or_404(
        CategoriaEvaluacion, id=cat_id, materia_nombre=materia
    )
    cat.delete()
    messages.success(request, "Categoría eliminada.")
    return redirect("profesor_categorias", cuat=cuat, materia=materia)


def profesor_cargar_notas(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    alumnos = _alumnos_por_materia(materia)
    categorias = _categorias_por_materia(materia)

    if not categorias:
        messages.warning(
            request,
            "Primero creá al menos una categoría de evaluación para esta materia.",
        )
        return redirect("profesor_categorias", cuat=cuat, materia=materia)

    if request.method == "POST":
        guardadas = 0
        for a in alumnos:
            for cat in categorias:
                key = f"celda_{a['id']}_{cat['id']}"
                val = request.POST.get(key, "").strip()

                if val == "":
                    NotaEvaluacion.objects.filter(
                        alumno_calificacion_id=a["id"],
                        categoria_id=cat["id"],
                        cuatrimestre=cuat,
                    ).delete()
                    continue

                try:
                    valor = Decimal(val)
                except InvalidOperation:
                    messages.warning(
                        request,
                        f"Valor inválido '{val}' para {a['nombre']} en {cat['nombre']}. Se omitió.",
                    )
                    continue

                if valor < 0 or valor > 10:
                    messages.warning(
                        request,
                        f"La nota '{val}' de {a['nombre']} en {cat['nombre']} está fuera del rango 0-10. Se omitió.",
                    )
                    continue

                NotaEvaluacion.objects.update_or_create(
                    alumno_calificacion_id=a["id"],
                    categoria_id=cat["id"],
                    cuatrimestre=cuat,
                    defaults={"valor": valor},
                )
                guardadas += 1

        messages.success(request, f"Se guardaron {guardadas} notas.")
        return redirect("profesor_ver_alumnos", cuat=cuat, materia=materia)

    notas_existentes = _notas_para_materia(materia, cuat)

    for a in alumnos:
        # celdas: lista de valores alineada con el orden de categorias
        celdas = []
        for cat in categorias:
            key = f"{a['id']}_{cat['id']}"
            lista = notas_existentes.get(key, [])
            celdas.append({
                "cat_id": cat["id"],
                "valor": lista[-1]["valor"] if lista else "",
            })
        a["celdas"] = celdas

    return render(request, "calificaciones/cargar_notas.html", {
        "cuat": cuat,
        "cuat_display": f"{cuat}° Cuatrimestre",
        "materia": materia,
        "alumnos": alumnos,
        "categorias": categorias,
    })


def profesor_cargar_excel(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    if request.method == "POST" and request.FILES.get("archivo"):
        try:
            import openpyxl
        except ImportError:
            messages.error(
                request,
                "La librería openpyxl no está instalada. Ejecutá: pip install openpyxl",
            )
            return redirect("profesor_cargar_excel", cuat=cuat, materia=materia)

        archivo = request.FILES["archivo"]
        wb = openpyxl.load_workbook(archivo, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        categorias = _categorias_por_materia(materia)
        cat_map = {c["nombre"]: c["id"] for c in categorias}

        rows_datos = []
        preview = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            nombre_alumno = str(row[0]).strip()
            alumno = AlumnoCalificacion.objects.filter(
                nombre__iexact=nombre_alumno, materia=materia
            ).first()

            fila = {
                "nombre": nombre_alumno,
                "encontrado": alumno is not None,
                "notas": {},
            }
            for i, header in enumerate(headers[1:], start=1):
                val = row[i] if i < len(row) else None
                if header in cat_map:
                    fila["notas"][header] = val
            preview.append(fila)

            row_dict = {
                "nombre": nombre_alumno,
                "alumno_id": alumno.id if alumno else None,
            }
            for header, cat_id in cat_map.items():
                row_dict[str(cat_id)] = None
            for i, header in enumerate(headers[1:], start=1):
                if header in cat_map:
                    val = row[i] if i < len(row) else None
                    row_dict[str(cat_map[header])] = str(val).strip() if val not in (None, "") else None
            rows_datos.append(row_dict)

        request.session[f"excel_{cuat}_{materia}"] = rows_datos
        return render(request, "calificaciones/confirmar_excel.html", {
            "cuat": cuat,
            "materia": materia,
            "preview": preview,
            "headers": headers,
        })

    return render(request, "calificaciones/cargar_excel.html", {
        "cuat": cuat,
        "materia": materia,
    })


def profesor_confirmar_excel(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    rows_datos = request.session.get(f"excel_{cuat}_{materia}")
    if not rows_datos:
        messages.error(request, "No se encontraron datos del Excel. Volvé a subir el archivo.")
        return redirect("profesor_cargar_excel", cuat=cuat, materia=materia)

    categorias = _categorias_por_materia(materia)
    cat_map = {c["id"]: c["nombre"] for c in categorias}

    guardadas = 0
    for row in rows_datos:
        if not row.get("alumno_id"):
            continue
        alumno = AlumnoCalificacion.objects.filter(id=row["alumno_id"], materia=materia).first()
        if not alumno:
            continue

        for cat in categorias:
            val = row.get(str(cat["id"]))
            if val is None:
                continue
            try:
                valor = Decimal(val)
            except (InvalidOperation, ValueError):
                continue

            if valor < 0 or valor > 10:
                continue

            NotaEvaluacion.objects.update_or_create(
                alumno_calificacion=alumno,
                categoria_id=cat["id"],
                cuatrimestre=cuat,
                defaults={"valor": valor},
            )
            guardadas += 1

    request.session.pop(f"excel_{cuat}_{materia}", None)
    messages.success(request, f"Se cargaron {guardadas} notas desde Excel.")
    return redirect("profesor_ver_alumnos", cuat=cuat, materia=materia)


def profesor_descargar_plantilla(request, cuat, materia):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        messages.error(request, "openpyxl no instalado.")
        return redirect("profesor_ver_alumnos", cuat=cuat, materia=materia)

    alumnos = _alumnos_por_materia(materia)
    categorias = _categorias_por_materia(materia)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{materia} - {cuat}° Cuat"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    ws.cell(row=1, column=1, value="Alumno")
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].border = thin_border
    ws["A1"].alignment = Alignment(horizontal="center")

    for i, cat in enumerate(categorias, start=2):
        cell = ws.cell(row=1, column=i, value=cat["nombre"])
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r, alumno in enumerate(alumnos, start=2):
        ws.cell(row=r, column=1, value=alumno["nombre"])
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, len(categorias) + 2):
            ws.cell(row=r, column=c).border = thin_border

    ws.column_dimensions["A"].width = 30
    for i in range(2, len(categorias) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"plantilla_{materia}_{cuat}cuat.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def profesor_toggle_boletin(request, cuat):
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    if request.method == "POST":
        config, _ = BoletinConfig.objects.get_or_create(cuatrimestre=cuat)
        config.publicado = not config.publicado
        if config.publicado:
            config.fecha_publicacion = timezone.now()
        config.save()

        estado = "publicado" if config.publicado else "ocultado"
        messages.success(request, f"Boletín {cuat}° cuatrimestre {estado}.")

    return redirect("calificaciones_menu")


# ─────────────────────────────────────────────
#  ALUMNO
# ─────────────────────────────────────────────

def alumno_seleccionar(request):
    nombres = (
        AlumnoCalificacion.objects.values_list("nombre", flat=True)
        .distinct()
        .order_by("nombre")
    )
    destino = request.GET.get("destino", "calificaciones")
    return render(request, "calificaciones/alumno_seleccionar.html", {
        "nombres": nombres,
        "destino": destino,
    })


def alumno_ver_calificaciones(request, cuat=None):
    if cuat is None:
        cuat = int(request.GET.get("cuat", 1))
    if cuat not in (1, 2):
        return redirect("calificaciones_menu")

    nombre = request.GET.get("nombre", "").strip()
    if not nombre:
        return redirect("alumno_seleccionar")

    alumnos = list(
        AlumnoCalificacion.objects.filter(nombre__iexact=nombre)
        .order_by("materia")
        .values("id", "materia")
    )

    materias_data = []
    for a in alumnos:
        cats = _categorias_por_materia(a["materia"])
        notas_qs = NotaEvaluacion.objects.filter(
            alumno_calificacion_id=a["id"],
            cuatrimestre=cuat,
        ).select_related("categoria")

        cats_data = []
        total = Decimal("0")
        count = 0
        for cat in cats:
            cat_notas = [
                {"valor": n.valor, "descripcion": n.descripcion}
                for n in notas_qs.filter(categoria_id=cat["id"])
            ]
            prom_cat = None
            if cat_notas:
                total_cat = sum((n["valor"] for n in cat_notas), Decimal("0"))
                prom_cat = _redondear(total_cat / Decimal(len(cat_notas)))
                total += prom_cat
                count += 1

            cats_data.append({
                "nombre": cat["nombre"],
                "notas": cat_notas,
                "promedio": prom_cat,
            })

        promedio = _redondear(total / Decimal(count)) if count > 0 else None
        materias_data.append({
            "materia": a["materia"],
            "categorias": cats_data,
            "promedio": promedio,
            "estado": _estado(promedio),
        })

    return render(request, "calificaciones/alumno_calificaciones.html", {
        "cuat": cuat,
        "cuat_display": f"{cuat}° Cuatrimestre",
        "nombre": nombre,
        "materias": materias_data,
    })


def alumno_ver_boletin(request, cuat=None):
    if cuat and cuat not in (1, 2):
        cuat = None

    nombre = request.GET.get("nombre", "").strip()
    if not nombre:
        return redirect("alumno_seleccionar")

    if cuat:
        configs = BoletinConfig.objects.filter(cuatrimestre=cuat, publicado=True)
    else:
        configs = BoletinConfig.objects.filter(publicado=True)

    if not configs.exists():
        messages.warning(request, "El boletín no está disponible aún.")
        return redirect("alumno_seleccionar")

    materias_data = []
    for config in configs:
        c = config.cuatrimestre
        alumnos = list(
            AlumnoCalificacion.objects.filter(nombre__iexact=nombre)
            .order_by("materia")
            .values("id", "materia")
        )

        for a in alumnos:
            prom = _promedio_alumno_cuat(a["id"], c)
            if prom is not None:
                materias_data.append({
                    "cuatrimestre": c,
                    "cuat_display": f"{c}° Cuatrimestre",
                    "materia": a["materia"],
                    "promedio": prom,
                    "estado": _estado(prom),
                })

    materias_data.sort(key=lambda x: (x["cuatrimestre"], x["materia"]))

    promedio_general = None
    if materias_data:
        total = sum((m["promedio"] for m in materias_data), Decimal("0"))
        promedio_general = _redondear(total / Decimal(len(materias_data)))

    return render(request, "calificaciones/boletin.html", {
        "nombre": nombre,
        "materias": materias_data,
        "promedio_general": promedio_general,
        "estado_general": _estado(promedio_general),
        "cuat_filtro": cuat,
    })
